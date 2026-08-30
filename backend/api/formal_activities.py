"""Customer-facing read-only API for standardized formal activity data."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import uuid

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from backend.auth.dependencies import get_current_user
from backend.database import get_db
from backend.models.user import User
from backend.services.formal_activity_ledger import (
    export_formal_activity_rows,
    get_formal_activity_detail,
    list_formal_activities,
)


router = APIRouter(prefix="/formal-activities", tags=["formal-activity-ledger"])


def _scope(user: User) -> tuple[uuid.UUID, uuid.UUID]:
    if user.tenant_id is None or user.enterprise_id is None:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="当前用户未绑定租户或企业",
        )
    return user.tenant_id, user.enterprise_id


def _filters(
    *,
    q: str | None,
    period_start: date | None,
    period_end: date | None,
    facility: str | None,
    category: str | None,
    document_type: str | None,
    calculation_status: str | None,
) -> dict:
    return {
        "query_text": q,
        "period_start": period_start,
        "period_end": period_end,
        "facility": facility,
        "category": category,
        "document_type": document_type,
        "calculation_status": calculation_status,
    }


@router.get("")
def list_ledger_records(
    q: str | None = Query(default=None, max_length=100),
    period_start: date | None = None,
    period_end: date | None = None,
    facility: str | None = Query(default=None, max_length=100),
    category: str | None = Query(default=None, max_length=50),
    document_type: str | None = Query(default=None, max_length=50),
    calculation_status: str | None = Query(default=None, alias="status", max_length=30),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    tenant_id, enterprise_id = _scope(user)
    try:
        return list_formal_activities(
            db,
            tenant_id=tenant_id,
            enterprise_id=enterprise_id,
            page=page,
            page_size=page_size,
            **_filters(
                q=q,
                period_start=period_start,
                period_end=period_end,
                facility=facility,
                category=category,
                document_type=document_type,
                calculation_status=calculation_status,
            ),
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.get("/export")
def export_ledger_records(
    q: str | None = Query(default=None, max_length=100),
    period_start: date | None = None,
    period_end: date | None = None,
    facility: str | None = Query(default=None, max_length=100),
    category: str | None = Query(default=None, max_length=50),
    document_type: str | None = Query(default=None, max_length=50),
    calculation_status: str | None = Query(default=None, alias="status", max_length=30),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    tenant_id, enterprise_id = _scope(user)
    try:
        rows = export_formal_activity_rows(
            db,
            tenant_id=tenant_id,
            enterprise_id=enterprise_id,
            **_filters(
                q=q,
                period_start=period_start,
                period_end=period_end,
                facility=facility,
                category=category,
                document_type=document_type,
                calculation_status=calculation_status,
            ),
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    output = _workbook(rows)
    filename = f"CarbonLab_standardized_data_ledger_{datetime.now():%Y%m%d}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{activity_id}")
def get_ledger_record(
    activity_id: uuid.UUID,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    tenant_id, enterprise_id = _scope(user)
    payload = get_formal_activity_detail(
        db,
        tenant_id=tenant_id,
        enterprise_id=enterprise_id,
        activity_id=activity_id,
    )
    if payload is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="标准化活动数据不存在或无权访问")
    return payload


def _workbook(rows: list[dict]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "标准化数据台账"
    headers = [
        "正式记录ID",
        "期间开始",
        "期间结束",
        "工厂",
        "活动类型",
        "数量",
        "单位",
        "原始文件",
        "原始文件哈希",
        "A-03质检状态",
        "自动检查覆盖得分",
        "H-01确认人",
        "确认时间",
        "核算状态",
        "排放结果",
        "排放单位",
        "版本",
        "正式记录哈希",
    ]
    sheet.append(headers)
    for item in rows:
        document = item.get("source_document") or {}
        result = item.get("emission_result") or {}
        sheet.append(
            [
                item["activity_data_id"],
                item["period"]["start"],
                item["period"]["end"],
                item["facility"]["name"],
                item["emission_source"]["category"],
                item["activity"]["quantity"],
                item["activity"]["unit"],
                document.get("filename"),
                document.get("content_hash"),
                item["quality"]["status"],
                item["quality"]["score"],
                item["confirmation"]["confirmed_by"],
                item["confirmation"]["confirmed_at"],
                item["calculation_status"],
                result.get("co2_tonnes"),
                result.get("unit"),
                item["confirmation"]["version"],
                item["content_hash"],
            ]
        )
    header_fill = PatternFill("solid", fgColor="EAF2FF")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="0B1736")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [38, 24, 24, 20, 24, 16, 12, 30, 68, 18, 20, 38, 28, 16, 16, 14, 10, 68]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[__import__("openpyxl").utils.get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
