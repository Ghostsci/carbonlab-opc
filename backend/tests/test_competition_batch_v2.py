"""End-to-end validation for 50-row synthetic electricity workbooks."""

from __future__ import annotations

from decimal import Decimal
import json
from pathlib import Path
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from passlib.context import CryptContext
import pytest

from backend.database import Base, get_engine, get_sessionmaker
from backend.main import app
from backend.models.emission_factor import EmissionFactor
from backend.models.enterprise import Enterprise
from backend.models.tenant import Tenant
from backend.models.user import User


ROOT = Path(__file__).resolve().parents[2]
DATASET_DIR = ROOT / "validation" / "competition_batch_v2"
MANIFEST = json.loads((DATASET_DIR / "manifest.json").read_text(encoding="utf-8"))
POSITIVE_CASES = MANIFEST["positive_cases"]
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


def setup_module() -> None:
    Base.metadata.create_all(bind=get_engine())


def teardown_module() -> None:
    Base.metadata.drop_all(bind=get_engine())


def _identity(case_code: str) -> tuple[User, EmissionFactor]:
    db = get_sessionmaker()()
    try:
        suffix = uuid.uuid4().hex[:12].upper()
        tenant = Tenant(name=f"batch-v2-{case_code}", slug=f"batch-v2-{case_code}-{suffix.lower()}")
        db.add(tenant)
        db.flush()
        enterprise = Enterprise(
            name=f"DEMO ONLY {case_code} 合成制造企业",
            unified_social_credit_code=f"92{suffix}0000",
            industry_code="C31",
            industry_name="黑色金属冶炼和压延加工业",
            tenant_id=tenant.id,
        )
        db.add(enterprise)
        db.flush()
        user = User(
            email=f"batch-v2-{case_code}-{suffix.lower()}@example.com",
            password_hash=pwd_context.hash("secret123"),
            role="admin",
            tenant_id=tenant.id,
            enterprise_id=enterprise.id,
        )
        factor = EmissionFactor(
            name="DEMO ONLY 华东电网合成测试因子",
            code=f"DEMO-BATCH-V2-{suffix}",
            category="electricity_grid",
            region="华东",
            year=2026,
            version_year=2026,
            is_default=True,
            value=Decimal("0.5"),
            unit="kgCO2e/kWh",
            source="synthetic-50-row-test-fixture",
        )
        db.add_all([user, factor])
        db.commit()
        db.refresh(user)
        db.refresh(factor)
        return user, factor
    finally:
        db.close()


def _headers(client: TestClient, user: User) -> dict[str, str]:
    login = client.post(
        "/api/auth/login",
        json={"email": user.email, "password": "secret123"},
    )
    assert login.status_code == 200, login.json()
    return {"Authorization": f"Bearer {login.json()['access_token']}"}


@pytest.mark.parametrize(
    "case",
    POSITIVE_CASES,
    ids=[Path(case["filename"]).stem for case in POSITIVE_CASES],
)
def test_each_workbook_contains_50_reconciling_detail_rows(case) -> None:
    """The user-visible workbook contains 50 rows and an exact summary reconciliation."""
    source = DATASET_DIR / case["filename"]
    workbook = load_workbook(source, read_only=False, data_only=False, keep_links=False)
    try:
        assert workbook.sheetnames == ["账单摘要", "50条用电明细", "数据说明"]
        detail = workbook["50条用电明细"]
        rows = list(detail.iter_rows(min_row=5, max_row=54, values_only=True))
        assert len(rows) == 50
        assert [row[0] for row in rows] == list(range(1, 51))
        assert all(row[1] and row[11] == "DEMO ONLY / SYNTHETIC" for row in rows)
        calculated_total = sum(
            (Decimal(str(row[6])) - Decimal(str(row[5]))) * Decimal(str(row[7]))
            for row in rows
        )
        assert calculated_total == Decimal(case["electricity_kwh"])
        assert detail["I55"].value == "=SUM(I5:I54)"
    finally:
        workbook.close()


@pytest.mark.parametrize(
    "case",
    POSITIVE_CASES,
    ids=[Path(case["filename"]).stem for case in POSITIVE_CASES],
)
def test_each_50_row_workbook_reaches_deterministic_calculation(case) -> None:
    """All 12 workbooks reach the controlled H-02/R-01 calculation path."""
    client = TestClient(app)
    user, factor = _identity(Path(case["filename"]).stem.lower())
    headers = _headers(client, user)
    source = DATASET_DIR / case["filename"]

    uploaded = client.post(
        "/api/upload",
        headers=headers,
        files={
            "file": (
                source.name,
                source.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert uploaded.status_code == 200, uploaded.json()
    document = uploaded.json()
    assert document["document_type"] == "electricity_bill"
    assert document["errors"] == []
    assert document["fields"]["period"] == case["period"]
    assert document["fields"]["facility"] == case["facility"]
    assert Decimal(document["fields"]["electricity_kwh"].split()[0].replace(",", "")) == Decimal(
        case["electricity_kwh"]
    )
    assert "50条用电明细" in document["raw_text"]
    assert "P050" in document["raw_text"]

    fields = {
        "electricity_kwh": case["electricity_kwh"],
        "period": case["period"],
        "facility": case["facility"],
    }
    candidate = client.post(
        f"/api/upload/{document['file_id']}/candidate",
        headers=headers,
        json={"fields": fields},
    )
    assert candidate.status_code == 200, candidate.json()
    quality = client.post(
        f"/api/upload/{document['file_id']}/quality-review",
        headers=headers,
        json={
            "candidate_token": candidate.json()["candidate_token"],
            "fields": fields,
        },
    )
    assert quality.status_code == 200, quality.json()
    assert quality.json()["quality_status"] in {"pass", "pass_with_warnings"}

    confirmed = client.post(
        "/api/upload/confirm-activity",
        headers=headers,
        json={
            "candidate_token": candidate.json()["candidate_token"],
            "quality_review_token": quality.json()["quality_review_token"],
            "file_id": document["file_id"],
            "document_content_hash": document["content_hash"],
            "filename": document["filename"],
            "document_type": "electricity_bill",
            "fields": fields,
        },
    )
    assert confirmed.status_code == 200, confirmed.json()
    formal = confirmed.json()["formal_write"]
    assert formal["calculation_status"] == "pending_factor"

    candidates = client.get(
        f"/api/upload/formal-activities/{formal['activity_data_id']}/factor-candidates",
        headers=headers,
    )
    assert candidates.status_code == 200, candidates.json()
    selected = next(
        item for item in candidates.json()["factor_candidates"] if item["factor_id"] == str(factor.id)
    )
    calculated = client.post(
        f"/api/upload/formal-activities/{formal['activity_data_id']}/confirm-factor",
        headers=headers,
        json={
            "factor_id": selected["factor_id"],
            "factor_snapshot_sha256": selected["factor_snapshot_sha256"],
            "selection_note": "人工核对50条合成明细、账期、区域、因子单位和来源后用于验证。",
        },
    )
    assert calculated.status_code == 200, calculated.json()
    result = calculated.json()["formal_write"]["emission_result"]
    assert Decimal(result["co2_tonnes_exact"]) == Decimal(
        case["expected_tco2e_at_demo_factor"]
    )
    assert result["unit"] == "tCO2e"
