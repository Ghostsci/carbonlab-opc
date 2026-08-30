"""Public API tests for the customer-facing standardized data ledger."""

from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
import uuid

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from passlib.context import CryptContext

from backend.database import Base, get_engine, get_sessionmaker
from backend.main import app
from backend.models.activity_data import ActivityData
from backend.models.document import DocumentStore
from backend.models.emission_source import EmissionSource
from backend.models.enterprise import Enterprise
from backend.models.site import Site
from backend.models.tenant import Tenant
from backend.models.user import User


pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


def setup_module():
    Base.metadata.create_all(bind=get_engine())


def teardown_module():
    Base.metadata.drop_all(bind=get_engine())


def _session():
    return get_sessionmaker()()


def _seed_activity(db, *, slug: str, email: str, quantity: str = "632600"):
    tenant = Tenant(name=slug.upper(), slug=slug)
    db.add(tenant)
    db.flush()
    enterprise = Enterprise(
        name=f"{slug} 制造企业",
        unified_social_credit_code=f"91{uuid.uuid4().hex[:16].upper()}",
        industry_code="C31",
        industry_name="黑色金属冶炼和压延加工业",
        tenant_id=tenant.id,
    )
    db.add(enterprise)
    db.flush()
    user = User(
        email=email,
        password_hash=pwd_context.hash("secret123"),
        role="admin",
        tenant_id=tenant.id,
        enterprise_id=enterprise.id,
    )
    site = Site(
        enterprise_id=enterprise.id,
        tenant_id=tenant.id,
        name="一号炼钢厂",
        address="江苏省苏州市",
        province="江苏",
        city="苏州",
        grid_region="华东",
    )
    db.add_all([user, site])
    db.flush()
    source = EmissionSource(
        site_id=site.id,
        tenant_id=tenant.id,
        name="一号炼钢厂 外购电力",
        scope="scope_2",
        category="purchased_electricity",
        fuel_type=None,
        source_code=f"TEST-{uuid.uuid4().hex[:12]}",
    )
    document = DocumentStore(
        tenant_id=tenant.id,
        enterprise_id=enterprise.id,
        filename="2026Q1_电费账单.xlsx",
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size_bytes=4096,
        storage_path=f"tests/{uuid.uuid4()}/electricity.xlsx",
        content_hash=uuid.uuid4().hex * 2,
        doc_type="electricity_bill",
        ocr_status="confirmed",
        ocr_result={
            "fields": {
                "electricity_kwh": f"{quantity} kWh",
                "period": "2026-Q1",
                "facility": "一号炼钢厂",
            },
            "field_sources": {
                "electricity_kwh": {
                    "kind": "spreadsheet_cell",
                    "sheet": "电费明细",
                    "cell": "E2",
                    "header_cell": "E1",
                    "header": "本期用电量",
                    "raw_value": quantity,
                    "unit": "kWh",
                }
            },
            "human_confirmation": {
                "actor_user_id": str(user.id),
                "candidate_id": "candidate-test",
                "quality_review": {
                    "quality_review_id": "quality-test",
                    "quality_status": "pass_with_warnings",
                    "score": 96,
                    "score_label": "自动检查覆盖得分",
                    "warnings_resolved": True,
                    "findings": [],
                    "resolutions": [],
                },
            },
        },
    )
    db.add_all([source, document])
    db.flush()
    activity = ActivityData(
        tenant_id=tenant.id,
        emission_source_id=source.id,
        period_start=datetime(2026, 1, 1, tzinfo=timezone.utc),
        period_end=datetime(2026, 3, 31, 23, 59, 59, tzinfo=timezone.utc),
        quantity=Decimal(quantity),
        unit="kWh",
        data_source="ocr",
        document_id=document.id,
        notes="source=data_inbox",
        derived_from=[f"document:{document.id}", "candidate:test"],
        content_hash=uuid.uuid4().hex * 2,
        idempotency_key=uuid.uuid4().hex * 2,
        version=1,
        confirmed_by=str(user.id),
        confirmed_at=datetime.now(timezone.utc),
    )
    db.add(activity)
    db.commit()
    db.refresh(activity)
    return tenant, enterprise, user, activity, document


def _login(client: TestClient, email: str) -> str:
    response = client.post("/api/auth/login", json={"email": email, "password": "secret123"})
    assert response.status_code == 200, response.json()
    return response.json()["access_token"]


def test_list_formal_activities_is_tenant_scoped_and_customer_readable():
    client = TestClient(app)
    db = _session()
    try:
        _, _, _, owned_activity, owned_document = _seed_activity(
            db,
            slug="ledger-owned",
            email="ledger-owned@example.com",
        )
        _seed_activity(
            db,
            slug="ledger-other",
            email="ledger-other@example.com",
            quantity="999999",
        )
        token = _login(client, "ledger-owned@example.com")

        response = client.get(
            "/api/formal-activities",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200, response.json()
        payload = response.json()
        assert payload["summary"] == {
            "total": 1,
            "calculated": 0,
            "pending_factor": 1,
            "source_documents": 1,
        }
        assert payload["pagination"] == {"page": 1, "page_size": 20, "total": 1, "pages": 1}
        assert len(payload["items"]) == 1
        item = payload["items"][0]
        assert item["activity_data_id"] == str(owned_activity.id)
        assert item["source_document"]["document_id"] == str(owned_document.id)
        assert item["source_document"]["filename"] == "2026Q1_电费账单.xlsx"
        assert item["facility"]["name"] == "一号炼钢厂"
        assert item["activity"]["quantity"] == "632600"
        assert item["activity"]["unit"] == "kWh"
        assert item["calculation_status"] == "pending_factor"
        assert item["quality"]["score"] == 96
        assert item["quality"]["warnings_resolved"] is True
    finally:
        db.close()


def test_formal_activity_detail_exposes_source_mapping_lineage_and_versions():
    client = TestClient(app)
    db = _session()
    try:
        _, _, _, activity, document = _seed_activity(
            db,
            slug="ledger-detail",
            email="ledger-detail@example.com",
        )
        token = _login(client, "ledger-detail@example.com")

        response = client.get(
            f"/api/formal-activities/{activity.id}",
            headers={"Authorization": f"Bearer {token}"},
        )

        assert response.status_code == 200, response.json()
        payload = response.json()
        assert payload["activity_data_id"] == str(activity.id)
        assert payload["formal_record"]["content_hash"] == activity.content_hash
        assert payload["formal_record"]["confirmed_by"]
        assert payload["source_document"]["download_url"] == f"/api/upload/{document.id}/download"
        electricity = next(
            item for item in payload["standardized_fields"] if item["canonical_key"] == "electricity_kwh"
        )
        assert electricity["raw_field"] == "electricity_kwh"
        assert electricity["raw_value"] == "632600 kWh"
        assert electricity["source_locator"]["cell"] == "E2"
        assert electricity["formal_destination"] == "ActivityData.quantity"
        assert payload["lineage"][0] == f"document:{document.id}"
        assert payload["version_history"][0]["version"] == 1
    finally:
        db.close()


def test_formal_activity_detail_hides_other_tenant_and_export_is_scoped():
    client = TestClient(app)
    db = _session()
    try:
        _, _, _, owned_activity, _ = _seed_activity(
            db,
            slug="ledger-export",
            email="ledger-export@example.com",
        )
        _, _, _, other_activity, _ = _seed_activity(
            db,
            slug="ledger-export-other",
            email="ledger-export-other@example.com",
            quantity="987654",
        )
        token = _login(client, "ledger-export@example.com")

        denied = client.get(
            f"/api/formal-activities/{other_activity.id}",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert denied.status_code == 404

        exported = client.get(
            "/api/formal-activities/export",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert exported.status_code == 200
        assert exported.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        workbook = load_workbook(filename=__import__("io").BytesIO(exported.content), read_only=True)
        rows = list(workbook["标准化数据台账"].iter_rows(values_only=True))
        assert len(rows) == 2
        assert str(owned_activity.id) in rows[1]
        assert str(other_activity.id) not in rows[1]
    finally:
        db.close()
