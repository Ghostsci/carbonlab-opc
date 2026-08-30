"""Capacity boundaries for locally parsed business workbooks."""

from pathlib import Path

from openpyxl import Workbook

from backend.ai.ocr import DocumentType, OCRService
from backend.services.digital_workforce import evaluate_document_quality


ROOT = Path(__file__).resolve().parents[2]
COMPETITION_1000_ROW_FIXTURE = (
    ROOT
    / "validation"
    / "competition_1000row"
    / "DEMO_ONLY_202601_HR_1000rows.xlsx"
)


def _write_1000_row_electricity_workbook(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "账单摘要"
    summary.append(
        [
            "供电公司",
            "户名",
            "户号",
            "账单月份",
            "本期用电量",
            "电价",
            "电费合计",
            "上期读数",
            "本期读数",
            "所属工厂",
        ]
    )
    summary.append(
        [
            "DEMO ONLY｜合成供电公司",
            "DEMO ONLY｜华盛钢铁演示企业",
            "DEMO-HR-2601",
            "2026年1月",
            2_480_000,
            0.718,
            1_780_640,
            184_200_000,
            186_680_000,
            "一号热轧生产装置（合成）",
        ]
    )

    details = workbook.create_sheet("1000条用电明细")
    details.append(
        [
            "序号",
            "抄表时间",
            "计量点编号",
            "班次",
            "期初读数",
            "期末读数",
            "本时段电量",
            "单位",
            "数据来源",
            "数据标识",
            "说明",
        ]
    )
    reading = 184_200_000
    for index in range(1, 1001):
        usage = 2_480
        next_reading = reading + usage
        details.append(
            [
                index,
                f"2026-01-{((index - 1) % 31) + 1:02d} 08:00:00",
                f"DEMO-METER-{((index - 1) % 12) + 1:02d}",
                ("早班", "中班", "晚班")[(index - 1) % 3],
                reading,
                next_reading,
                usage,
                "kWh",
                "企业能源管理系统导出的逐时计量记录（合成演示数据）",
                "DEMO ONLY / SYNTHETIC / NOT FOR REPORTING",
                "用于验证千行结构化电力明细的本地解析、检索和安全容量边界",
            ]
        )
        reading = next_reading

    notes = workbook.create_sheet("数据说明")
    notes.append(["属性", "说明"])
    notes.append(["数据性质", "完全合成，仅用于测试，不代表真实客户或监管申报数据"])
    notes.append(["验证目标", "验证1000行结构化XLSX能够被安全读取并提取账单摘要"])
    workbook.save(path)
    workbook.close()


def test_ocr_accepts_governed_1000_row_xlsx_and_extracts_summary(tmp_path: Path):
    workbook_path = tmp_path / "DEMO_ONLY_1000_rows.xlsx"
    _write_1000_row_electricity_workbook(workbook_path)

    result = OCRService(use_llm_fallback=False).process(workbook_path)

    assert result.read_status == "read", result.errors
    assert result.document_type == DocumentType.ELECTRICITY_BILL
    assert result.fields["electricity_kwh"] == "2480000"
    assert result.fields["period"] == "2026年1月"
    assert len(result.raw_text) > 100_000


def test_ocr_still_rejects_xlsx_beyond_its_configured_text_boundary(tmp_path: Path):
    workbook_path = tmp_path / "DEMO_ONLY_over_limit.xlsx"
    _write_1000_row_electricity_workbook(workbook_path)
    service = OCRService(use_llm_fallback=False)
    service.MAX_XLSX_EXTRACTED_CHARS = 50_000

    result = service.process(workbook_path)

    assert result.read_status == "reject"
    assert result.document_type == DocumentType.UNKNOWN
    assert result.raw_text == ""
    assert result.errors == ["XLSX extracted text exceeds the processing limit."]


def test_packaged_1000_row_competition_fixture_remains_processable():
    result = OCRService(use_llm_fallback=False).process(COMPETITION_1000_ROW_FIXTURE)

    assert result.read_status == "read", result.errors
    assert result.document_type == DocumentType.ELECTRICITY_BILL
    assert result.fields["electricity_kwh"] == "2480000"
    assert result.fields["meter_reading_start"] == "184200000"
    assert result.fields["meter_reading_end"] == "186680000"
    assert "1000条用电明细" in result.raw_text

    electricity_source = result.field_sources["electricity_kwh"]
    assert electricity_source["sheet"] == "账单摘要"
    assert electricity_source["cell"] == "E5"
    assert electricity_source["header_cell"] == "E4"
    assert electricity_source["unit"] == "kWh"
    assert electricity_source["unit_source"] == "header"

    quality = evaluate_document_quality(
        document_type=result.document_type.value,
        document_content_hash="a" * 64,
        fields=result.fields,
        source_snapshot={
            "fields": result.fields,
            "raw_text": result.raw_text,
            "field_sources": result.field_sources,
        },
        retrieval_evidence={
            key: {
                "retrieval_run_id": f"run-{key}",
                "hits": [{"field_keys": [key], "excerpt": result.field_sources[key]["excerpt"]}],
            }
            for key in ("electricity_kwh", "period", "facility")
        },
    )
    unit_finding = next(item for item in quality["findings"] if item["check_key"] == "quantity_unit")
    assert unit_finding["result"] == "pass"
    assert unit_finding["source_locator"]["cell"] == "E5"
    assert "kWh" in unit_finding["message"]
