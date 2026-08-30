"""Fail-closed document text reading and structured field extraction."""

from __future__ import annotations

import csv
import datetime as dt
import io
import math
import re
import unicodedata
import zipfile
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any


class DocumentType(str, Enum):
    INVOICE = "invoice"
    ELECTRICITY_BILL = "electricity_bill"
    PRODUCTION_REPORT = "production_report"
    UNKNOWN = "unknown"


@dataclass
class OCRResult:
    document_type: DocumentType
    fields: dict[str, str]
    confidence: float
    raw_text: str
    tables: list[dict[str, Any]] = field(default_factory=list)
    field_sources: dict[str, dict[str, Any]] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)
    read_status: str = "read"
    reader: str = "unknown"


class DocumentReadError(ValueError):
    """The supplied file is malformed, unsafe, or not valid for its suffix."""


class DocumentReaderUnavailable(ValueError):
    """The file may be valid, but no safe reader is configured for it."""


class OCRService:
    """Read supported business documents without inventing OCR output.

    UTF-8 text, text-layer PDFs, and XLSX workbooks are read locally. Image
    documents deliberately abstain until a real OCR adapter is configured;
    binary placeholder text is never returned as document content.
    """

    MAX_FILE_BYTES = 10 * 1024 * 1024
    # Text-layer PDFs and structured workbooks have different density and
    # memory profiles. Keep the PDF boundary conservative while allowing a
    # governed 1,000-row business workbook without removing XLSX safeguards.
    MAX_PDF_EXTRACTED_CHARS = 100_000
    MAX_XLSX_EXTRACTED_CHARS = 250_000
    MAX_PDF_PAGES = 50
    MAX_XLSX_SHEETS = 20
    MAX_XLSX_ROWS_PER_SHEET = 5_000
    MAX_XLSX_COLUMNS = 100
    MAX_XLSX_ARCHIVE_MEMBERS = 2_000
    MAX_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024

    SOURCE_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
        "supplier_name": ("供电公司", "供应商", "供电单位"),
        "customer_name": ("户名", "用户名称", "客户名称"),
        "customer_number": ("户号", "客户编号", "用户编号"),
        "period": ("账单月份", "报告期间", "统计期间", "计费周期", "抄表日期"),
        "electricity_kwh": ("本期用电量", "用电量", "有功电量", "总电量"),
        "unit_price": ("电价", "单价"),
        "total_amount": ("电费合计", "应收电费", "合计金额", "金额"),
        "meter_reading_start": ("上期读数", "上次读数", "期初读数"),
        "meter_reading_end": ("本期读数", "本次读数", "期末读数"),
        "facility": ("所属工厂", "所属设施", "工厂", "生产区域", "用电地址"),
    }

    def __init__(self, use_llm_fallback: bool = True):
        # Kept for API compatibility. Unsafe implicit LLM/OCR fallback is not
        # used; callers must configure a real image OCR adapter explicitly.
        self.use_llm_fallback = use_llm_fallback

    def process(
        self,
        file_path: str | Path,
        doc_type_hint: DocumentType | None = None,
    ) -> OCRResult:
        path = Path(file_path)
        if not path.exists():
            return self._failed_result(
                f"File not found: {file_path}",
                read_status="reject",
                reader="failed",
            )
        if not path.is_file():
            return self._failed_result(
                "Document path is not a regular file.",
                read_status="reject",
                reader="failed",
            )

        try:
            text, reader = self._extract_text(path)
        except DocumentReaderUnavailable as exc:
            return self._failed_result(
                str(exc),
                read_status="abstain",
                reader="unavailable",
            )
        except DocumentReadError as exc:
            return self._failed_result(
                str(exc),
                read_status="reject",
                reader="failed",
            )
        except OSError:
            return self._failed_result(
                "Document could not be read.",
                read_status="reject",
                reader="failed",
            )

        if not text.strip():
            return self._failed_result(
                "No text extracted from document.",
                read_status="abstain",
                reader=reader,
            )

        doc_type = doc_type_hint or self._detect_type(text)
        if doc_type == DocumentType.UNKNOWN:
            return OCRResult(
                document_type=DocumentType.UNKNOWN,
                fields={},
                confidence=0,
                raw_text=text,
                errors=["Document type could not be determined; no fields were extracted."],
                read_status="abstain",
                reader=reader,
            )

        try:
            fields = extract_fields(text, doc_type)
        except ValueError as exc:
            return OCRResult(
                document_type=doc_type,
                fields={},
                confidence=0,
                raw_text=text,
                errors=[str(exc)],
                read_status="abstain",
                reader=reader,
            )
        except Exception:
            return OCRResult(
                document_type=doc_type,
                fields={},
                confidence=0,
                raw_text=text,
                errors=["Structured field extraction failed."],
                read_status="reject",
                reader=reader,
            )
        confidence = self._estimate_confidence(text, fields, doc_type)
        field_sources = self._locate_field_sources(path, text, fields)
        if not any(value.strip() for value in fields.values()):
            return OCRResult(
                document_type=doc_type,
                fields=fields,
                confidence=0,
                raw_text=text,
                errors=["No supported structured fields were extracted."],
                read_status="abstain",
                reader=reader,
            )

        return OCRResult(
            document_type=doc_type,
            fields=fields,
            confidence=round(confidence, 1),
            raw_text=text,
            tables=self._extract_tables(text),
            field_sources=field_sources,
            read_status="read",
            reader=reader,
        )

    def _failed_result(
        self,
        error: str,
        *,
        read_status: str,
        reader: str,
    ) -> OCRResult:
        return OCRResult(
            document_type=DocumentType.UNKNOWN,
            fields={},
            confidence=0,
            raw_text="",
            errors=[error],
            read_status=read_status,
            reader=reader,
        )

    def _extract_text(self, path: Path) -> tuple[str, str]:
        try:
            file_size = path.stat().st_size
        except OSError as exc:
            raise DocumentReadError("Document metadata could not be read.") from exc
        if file_size > self.MAX_FILE_BYTES:
            raise DocumentReadError("Document exceeds the 10 MB processing limit.")

        suffix = path.suffix.lower()
        if suffix in {".txt", ".csv"}:
            try:
                content = path.read_bytes()
            except OSError as exc:
                raise DocumentReadError("Text document could not be read.") from exc
            return self._decode_text(content), "text"
        if suffix == ".pdf":
            return self._extract_pdf(path), "pypdf"
        if suffix == ".xlsx":
            return self._extract_xlsx(path), "openpyxl"
        if suffix == ".xls":
            raise DocumentReaderUnavailable(
                "Legacy .xls reader is not configured; convert the file to .xlsx or CSV."
            )
        if suffix in {".png", ".jpg", ".jpeg"}:
            raise DocumentReaderUnavailable(
                "Image OCR is not configured; the document requires OCR or human review."
            )
        raise DocumentReaderUnavailable(
            f"No document reader is configured for {suffix or 'this file type'}."
        )

    def _decode_text(self, content: bytes) -> str:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise DocumentReadError("Text document is not valid UTF-8.") from exc
        self._validate_text(text, source="Text document")
        return text

    def _validate_text(self, text: str, *, source: str) -> None:
        if "\x00" in text:
            raise DocumentReadError(f"{source} contains binary NUL bytes.")
        control_count = sum(
            1
            for character in text
            if unicodedata.category(character) == "Cc"
            and character not in {"\n", "\r", "\t"}
        )
        if control_count:
            raise DocumentReadError(f"{source} contains binary control characters.")

    def _extract_pdf(self, path: Path) -> str:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise DocumentReaderUnavailable(
                "PDF text reader is not installed; pypdf is required."
            ) from exc

        try:
            reader = PdfReader(str(path), strict=True)
            if reader.is_encrypted:
                raise DocumentReadError("Encrypted PDF cannot be processed.")
            if len(reader.pages) > self.MAX_PDF_PAGES:
                raise DocumentReadError(
                    f"PDF exceeds the {self.MAX_PDF_PAGES}-page processing limit."
                )

            page_texts: list[str] = []
            extracted_chars = 0
            for page in reader.pages:
                page_text = page.extract_text() or ""
                extracted_chars += len(page_text) + 1
                if extracted_chars > self.MAX_PDF_EXTRACTED_CHARS:
                    raise DocumentReadError("PDF extracted text exceeds the processing limit.")
                page_texts.append(page_text)
        except (DocumentReadError, DocumentReaderUnavailable):
            raise
        except Exception as exc:
            raise DocumentReadError("PDF is malformed or unreadable.") from exc

        text = "\n".join(page_texts)
        if not text.strip():
            raise DocumentReaderUnavailable(
                "PDF has no readable text layer; image OCR or human review is required."
            )
        self._validate_text(text, source="PDF text layer")
        return text

    def _extract_xlsx(self, path: Path) -> str:
        self._validate_xlsx_archive(path)
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise DocumentReaderUnavailable(
                "XLSX reader is not installed; openpyxl is required."
            ) from exc

        try:
            workbook = load_workbook(
                filename=path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as exc:
            raise DocumentReadError("XLSX workbook is malformed or unreadable.") from exc

        lines: list[str] = []
        extracted_chars = 0
        try:
            if len(workbook.worksheets) > self.MAX_XLSX_SHEETS:
                raise DocumentReadError(
                    f"XLSX exceeds the {self.MAX_XLSX_SHEETS}-sheet processing limit."
                )

            for worksheet in workbook.worksheets:
                if worksheet.max_row and worksheet.max_row > self.MAX_XLSX_ROWS_PER_SHEET:
                    raise DocumentReadError(
                        "XLSX exceeds the row-per-sheet processing limit."
                    )
                if worksheet.max_column and worksheet.max_column > self.MAX_XLSX_COLUMNS:
                    raise DocumentReadError(
                        "XLSX exceeds the column-per-sheet processing limit."
                    )

                sheet_lines: list[str] = []
                for row_index, row in enumerate(
                    worksheet.iter_rows(values_only=True),
                    start=1,
                ):
                    if row_index > self.MAX_XLSX_ROWS_PER_SHEET:
                        raise DocumentReadError(
                            "XLSX exceeds the row-per-sheet processing limit."
                        )
                    if len(row) > self.MAX_XLSX_COLUMNS:
                        raise DocumentReadError(
                            "XLSX exceeds the column-per-sheet processing limit."
                        )
                    cells = [self._cell_text(value) for value in row]
                    if not any(cells):
                        continue
                    line = self._csv_line(cells)
                    sheet_lines.append(line)
                    extracted_chars += len(line) + 1
                    if extracted_chars > self.MAX_XLSX_EXTRACTED_CHARS:
                        raise DocumentReadError(
                            "XLSX extracted text exceeds the processing limit."
                        )

                if sheet_lines:
                    title = str(worksheet.title).replace("\r", " ").replace("\n", " ")
                    heading = f"[Sheet: {title}]"
                    lines.append(heading)
                    lines.extend(sheet_lines)
                    extracted_chars += len(heading) + 1
                    if extracted_chars > self.MAX_XLSX_EXTRACTED_CHARS:
                        raise DocumentReadError(
                            "XLSX extracted text exceeds the processing limit."
                        )
        finally:
            workbook.close()

        text = "\n".join(lines)
        self._validate_text(text, source="XLSX content")
        return text

    def _validate_xlsx_archive(self, path: Path) -> None:
        try:
            with zipfile.ZipFile(path) as archive:
                members = archive.infolist()
                names = {member.filename for member in members}
                if len(members) > self.MAX_XLSX_ARCHIVE_MEMBERS:
                    raise DocumentReadError("XLSX archive contains too many members.")
                if any(member.flag_bits & 0x1 for member in members):
                    raise DocumentReadError("Encrypted XLSX workbooks cannot be processed.")
                if sum(member.file_size for member in members) > self.MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise DocumentReadError("XLSX archive expands beyond the processing limit.")
                required_members = {"[Content_Types].xml", "xl/workbook.xml"}
                if not required_members.issubset(names):
                    raise DocumentReadError("XLSX archive is missing required workbook data.")
        except DocumentReadError:
            raise
        except (OSError, zipfile.BadZipFile) as exc:
            raise DocumentReadError("XLSX workbook is malformed or unreadable.") from exc

    def _cell_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, dt.datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, dt.date):
            return value.isoformat()
        if isinstance(value, float):
            return format(value, ".15g") if math.isfinite(value) else ""
        return str(value).replace("\r", " ").replace("\n", " ")

    def _csv_line(self, cells: list[str]) -> str:
        output = io.StringIO()
        csv.writer(output, lineterminator="").writerow(cells)
        return output.getvalue()

    def _detect_type(self, text: str) -> DocumentType:
        normalized = text.casefold()
        signals = {
            DocumentType.INVOICE: ("发票代码", "发票号码", "价税合计", "invoice"),
            DocumentType.ELECTRICITY_BILL: (
                "电费",
                "用电量",
                "有功电量",
                "总电量",
                "kwh",
                "kw·h",
                "供电公司",
            ),
            DocumentType.PRODUCTION_REPORT: (
                "生产报表",
                "生产报告",
                "合格产量",
                "生产产量",
                "产量",
                "产出",
            ),
        }
        scores = {
            document_type: sum(keyword in normalized for keyword in keywords)
            for document_type, keywords in signals.items()
        }
        best_type = max(scores, key=scores.get)
        return best_type if scores[best_type] else DocumentType.UNKNOWN

    def _estimate_confidence(
        self,
        text: str,
        fields: dict[str, str],
        doc_type: DocumentType | None = None,
    ) -> float:
        del text  # Reserved for a future reader-quality signal.
        if not fields:
            return 0.0
        populated = sum(1 for value in fields.values() if value.strip())
        if not populated:
            return 0.0

        confidence = min(95.0, (populated / len(fields)) * 100)
        if doc_type == DocumentType.ELECTRICITY_BILL:
            if not fields.get("electricity_kwh"):
                confidence = min(confidence, 40.0)
            elif not fields.get("period"):
                confidence = min(confidence, 60.0)
        elif doc_type == DocumentType.PRODUCTION_REPORT:
            if not fields.get("production_output"):
                confidence = min(confidence, 40.0)
        return confidence

    def _extract_tables(self, text: str) -> list[dict[str, Any]]:
        """Reserved for a future layout-aware table adapter."""
        del text
        return []

    def _locate_field_sources(
        self,
        path: Path,
        text: str,
        fields: dict[str, str],
    ) -> dict[str, dict[str, Any]]:
        """Locate extracted candidates in the owned source document.

        The locator is deliberately compact: only extracted business fields are
        indexed, rather than persisting every cell from a large workbook.  It is
        sufficient for a reviewer to jump from an A-03 finding back to the exact
        source row/cell and see the unit context that produced the candidate.
        """
        try:
            if path.suffix.lower() == ".xlsx":
                return self._locate_xlsx_field_sources(path, fields)
            if path.suffix.lower() in {".csv", ".txt"}:
                return self._locate_delimited_field_sources(text, fields)
        except (OSError, ValueError, csv.Error):
            # Location metadata improves reviewability but must never turn a
            # readable document into a failed upload.  A-03 will fall back to a
            # text-line locator and require explicit human confirmation.
            pass
        return self._locate_text_field_sources(text, fields)

    def _locate_xlsx_field_sources(
        self,
        path: Path,
        fields: dict[str, str],
    ) -> dict[str, dict[str, Any]]:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        located: dict[str, dict[str, Any]] = {}
        raw_line = 0
        try:
            for worksheet in workbook.worksheets:
                raw_line += 1  # [Sheet: ...] marker emitted by _extract_xlsx.
                active_headers: dict[str, tuple[int, str, int]] = {}
                unit_column: int | None = None
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    cells = [self._cell_text(value) for value in row]
                    if not any(cells):
                        continue
                    raw_line += 1
                    detected_headers: dict[str, tuple[int, str, int]] = {}
                    detected_unit_column: int | None = None
                    for column_index, cell in enumerate(cells, start=1):
                        key = self._source_field_key(cell)
                        if key:
                            detected_headers[key] = (column_index, cell, row_index)
                        if self._normalize_source_label(cell) in {"单位", "计量单位"}:
                            detected_unit_column = column_index
                    if detected_headers:
                        active_headers = detected_headers
                        unit_column = detected_unit_column
                        continue

                    for field_key, candidate in fields.items():
                        if field_key in located or candidate in (None, ""):
                            continue
                        header = active_headers.get(field_key)
                        if header and header[0] <= len(cells):
                            column_index, header_text, header_row = header
                            raw_value = cells[column_index - 1]
                            if self._source_values_match(candidate, raw_value):
                                unit, unit_source = self._source_unit_context(
                                    header_text=header_text,
                                    raw_value=raw_value,
                                    unit_value=(
                                        cells[unit_column - 1]
                                        if unit_column and unit_column <= len(cells)
                                        else ""
                                    ),
                                )
                                column_letter = get_column_letter(column_index)
                                located[field_key] = {
                                    "kind": "spreadsheet_cell",
                                    "sheet": worksheet.title,
                                    "row": row_index,
                                    "column": column_index,
                                    "column_label": column_letter,
                                    "cell": f"{column_letter}{row_index}",
                                    "header_cell": f"{column_letter}{header_row}",
                                    "header": header_text,
                                    "raw_value": raw_value,
                                    "unit": unit,
                                    "unit_source": unit_source,
                                    "text_line_start": raw_line,
                                    "text_line_end": raw_line,
                                    "excerpt": self._row_excerpt(cells, column_index),
                                }
                                continue

                        # Also support label/value layouts such as
                        # "本期用电量 | 632600 | kWh".
                        for column_index, label in enumerate(cells, start=1):
                            if self._source_field_key(label) != field_key:
                                continue
                            for value_column in range(column_index + 1, min(len(cells), column_index + 3) + 1):
                                raw_value = cells[value_column - 1]
                                if not self._source_values_match(candidate, raw_value):
                                    continue
                                unit_value = cells[value_column] if value_column < len(cells) else ""
                                unit, unit_source = self._source_unit_context(
                                    header_text=label,
                                    raw_value=raw_value,
                                    unit_value=unit_value,
                                )
                                column_letter = get_column_letter(value_column)
                                located[field_key] = {
                                    "kind": "spreadsheet_cell",
                                    "sheet": worksheet.title,
                                    "row": row_index,
                                    "column": value_column,
                                    "column_label": column_letter,
                                    "cell": f"{column_letter}{row_index}",
                                    "header_cell": f"{get_column_letter(column_index)}{row_index}",
                                    "header": label,
                                    "raw_value": raw_value,
                                    "unit": unit,
                                    "unit_source": unit_source,
                                    "text_line_start": raw_line,
                                    "text_line_end": raw_line,
                                    "excerpt": self._row_excerpt(cells, value_column),
                                }
                                break
            return located
        finally:
            workbook.close()

    def _locate_delimited_field_sources(
        self,
        text: str,
        fields: dict[str, str],
    ) -> dict[str, dict[str, Any]]:
        located: dict[str, dict[str, Any]] = {}
        rows = list(csv.reader(io.StringIO(text)))
        active_headers: dict[str, tuple[int, str, int]] = {}
        unit_column: int | None = None
        for row_index, row in enumerate(rows, start=1):
            cells = [str(value).strip() for value in row]
            detected_headers: dict[str, tuple[int, str, int]] = {}
            detected_unit_column: int | None = None
            for column_index, cell in enumerate(cells, start=1):
                key = self._source_field_key(cell)
                if key:
                    detected_headers[key] = (column_index, cell, row_index)
                if self._normalize_source_label(cell) in {"单位", "计量单位"}:
                    detected_unit_column = column_index
            if detected_headers:
                active_headers = detected_headers
                unit_column = detected_unit_column
                continue
            for field_key, candidate in fields.items():
                if field_key in located or candidate in (None, ""):
                    continue
                header = active_headers.get(field_key)
                if not header or header[0] > len(cells):
                    continue
                column_index, header_text, header_row = header
                raw_value = cells[column_index - 1]
                if not self._source_values_match(candidate, raw_value):
                    continue
                unit, unit_source = self._source_unit_context(
                    header_text=header_text,
                    raw_value=raw_value,
                    unit_value=(cells[unit_column - 1] if unit_column and unit_column <= len(cells) else ""),
                )
                located[field_key] = {
                    "kind": "delimited_cell",
                    "row": row_index,
                    "column": column_index,
                    "cell": f"R{row_index}C{column_index}",
                    "header_cell": f"R{header_row}C{column_index}",
                    "header": header_text,
                    "raw_value": raw_value,
                    "unit": unit,
                    "unit_source": unit_source,
                    "text_line_start": row_index,
                    "text_line_end": row_index,
                    "excerpt": self._row_excerpt(cells, column_index),
                }
        return located

    def _locate_text_field_sources(
        self,
        text: str,
        fields: dict[str, str],
    ) -> dict[str, dict[str, Any]]:
        located: dict[str, dict[str, Any]] = {}
        for field_key, candidate in fields.items():
            if candidate in (None, ""):
                continue
            for line_number, line in enumerate(text.splitlines(), start=1):
                if not self._source_values_match(candidate, line):
                    continue
                unit, unit_source = self._source_unit_context(
                    header_text=line,
                    raw_value=line,
                    unit_value="",
                )
                located[field_key] = {
                    "kind": "text_line",
                    "text_line_start": line_number,
                    "text_line_end": line_number,
                    "raw_value": candidate,
                    "unit": unit,
                    "unit_source": unit_source,
                    "excerpt": line[:500],
                }
                break
        return located

    def _source_field_key(self, value: Any) -> str | None:
        normalized = self._normalize_source_label(value)
        if not normalized:
            return None
        for field_key, aliases in self.SOURCE_HEADER_ALIASES.items():
            for alias in aliases:
                alias_normalized = self._normalize_source_label(alias)
                if normalized == alias_normalized or normalized.startswith(alias_normalized):
                    return field_key
        return None

    def _normalize_source_label(self, value: Any) -> str:
        normalized = re.sub(r"\s+", "", str(value or "").strip().lower())
        normalized = normalized.replace("（", "(").replace("）", ")")
        return re.sub(r"\([^)]*\)$", "", normalized)

    def _normalize_source_value(self, value: Any) -> str:
        text = str(value or "").strip().lower().replace(",", "")
        return re.sub(r"[^0-9a-z\u4e00-\u9fff.+-]+", "", text)

    def _source_values_match(self, candidate: Any, raw_value: Any) -> bool:
        needle = self._normalize_source_value(candidate)
        haystack = self._normalize_source_value(raw_value)
        if not needle or not haystack:
            return False
        return needle == haystack or needle in haystack or haystack in needle

    def _source_unit_context(
        self,
        *,
        header_text: str,
        raw_value: str,
        unit_value: str,
    ) -> tuple[str | None, str | None]:
        unit_patterns = (
            (r"(?i)kw[·.]?h|千瓦时", "kWh"),
            (r"(?i)mwh|兆瓦时", "MWh"),
            (r"(?i)wh|瓦时", "Wh"),
            (r"(?i)gj", "GJ"),
            (r"(?i)mj", "MJ"),
        )
        for source_name, value in (
            ("value", raw_value),
            ("header", header_text),
            ("unit_column", unit_value),
        ):
            for pattern, canonical in unit_patterns:
                if re.search(pattern, str(value or "")):
                    return canonical, source_name
        return None, None

    def _row_excerpt(self, cells: list[str], focus_column: int) -> str:
        start = max(0, focus_column - 3)
        end = min(len(cells), focus_column + 2)
        return " | ".join(cell for cell in cells[start:end] if cell)[:500]


_ocr_service = OCRService()


def extract_fields(text: str, doc_type: DocumentType) -> dict[str, str]:
    from backend.ai.extractors import (
        ElectricityBillExtractor,
        InvoiceExtractor,
        ProductionReportExtractor,
    )

    if doc_type == DocumentType.INVOICE:
        return InvoiceExtractor().extract(text)
    if doc_type == DocumentType.ELECTRICITY_BILL:
        return ElectricityBillExtractor().extract(text)
    if doc_type == DocumentType.PRODUCTION_REPORT:
        return ProductionReportExtractor().extract(text)
    return {}
